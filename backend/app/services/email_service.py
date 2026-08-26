"""
Email Service — Automated Supplier Notification with Rate-Limiting, No-Reply Protocol,
Ownership Locking, and Batch Delivery.
"""

import asyncio
import logging
import smtplib
from datetime import datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
import secrets

from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.master import SupplierMaster
from app.models.po import POHeader, POItem, POItemAuditLog
from app.models.supplier_token import SupplierPortalToken
from app.models.system_setting import SystemSetting

logger = logging.getLogger(__name__)


def calculate_prd_expiry_date(now_dt: datetime) -> datetime:
    """
    PRD Expiration Rules:
    - Monday (0), Tuesday (1), Wednesday (2) -> Expires Wednesday 23:59:59
    - Thursday (3), Friday (4), Saturday (5), Sunday (6) -> Expires Sunday 23:59:59
    """
    weekday = now_dt.weekday()
    if weekday <= 2:
        days_ahead = 2 - weekday
        return (now_dt + timedelta(days=days_ahead)).replace(hour=23, minute=59, second=59, microsecond=0)
    else:
        days_ahead = 6 - weekday
        return (now_dt + timedelta(days=days_ahead)).replace(hour=23, minute=59, second=59, microsecond=0)


async def get_or_create_supplier_token(db: AsyncSession, supplier_code: str) -> SupplierPortalToken:
    """
    Invalidate all previous active tokens for this supplier and create a fresh cryptographic token.
    Enforces that opening an older email link will be rejected / expired immediately.
    """
    now_dt = datetime.now(timezone.utc)
    target_expiry = calculate_prd_expiry_date(now_dt)

    # 1. Invalidate / Revoke all old active tokens for this supplier
    stmt_revoke = (
        select(SupplierPortalToken)
        .where(SupplierPortalToken.supplier_code == supplier_code)
        .where(SupplierPortalToken.is_submitted == False)
    )
    old_tokens = (await db.execute(stmt_revoke)).scalars().all()
    for old_tok in old_tokens:
        old_tok.expires_at = now_dt - timedelta(seconds=1)
        old_tok.is_submitted = True
        db.add(old_tok)

    # 2. Generate a brand new cryptographic token
    raw_token = f"tok_{secrets.token_hex(20)}"
    token_obj = SupplierPortalToken(
        token=raw_token,
        supplier_code=supplier_code,
        po_number=None,
        is_submitted=False,
        expires_at=target_expiry,
    )
    db.add(token_obj)
    await db.commit()
    await db.refresh(token_obj)

    return token_obj


def build_no_reply_email_html(
    supplier_name: str,
    portal_url: str,
    expiry_formatted: str,
    open_po_count: int,
) -> str:
    """Builds HTML email body with clear No-Reply header and instructions."""
    return f"""
    <!DOCTYPE html>
    <html>
      <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
      </head>
      <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #1e293b; background-color: #f8fafc; margin: 0; padding: 24px; line-height: 1.6;">
        <div style="max-width: 600px; margin: 0 auto; background: #ffffff; border-radius: 16px; border: 1px solid #e2e8f0; overflow: hidden; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);">
          <!-- Header Banner -->
          <div style="background: linear-gradient(135deg, #0284c7, #0369a1); padding: 24px 32px; color: #ffffff;">
            <h1 style="margin: 0; font-size: 20px; font-weight: bold; letter-spacing: -0.5px;">ระบบติดตามการรับวัตถุดิบ (IRM System)</h1>
            <p style="margin: 4px 0 0 0; font-size: 13px; opacity: 0.9;">ฝ่ายจัดซื้อ — แจ้งระบุวันและจำนวนส่งมอบวัตถุดิบ</p>
          </div>

          <!-- Content Body -->
          <div style="padding: 32px;">
            <p style="font-size: 15px; margin-top: 0;">เรียน <strong>{supplier_name}</strong>,</p>
            <p style="font-size: 14px; color: #475569;">
              ฝ่ายจัดซื้อ Window Asia PCL. ขอเรียนให้ท่านเข้าสู่ระบบเพื่อระบุวันและจำนวนส่งมอบสินค้า สำหรับรายการสั่งซื้อ (PO) ที่เปิดค้างอยู่ทั้งหมด <strong>{open_po_count}</strong> รายการ
            </p>

            <!-- Important Expiry Alert -->
            <div style="background-color: #fffbeb; border: 1px solid #fef08a; border-radius: 12px; padding: 14px 18px; margin: 20px 0;">
              <div style="color: #854d0e; font-size: 13px; font-weight: bold; margin-bottom: 4px;">⏰ รอบเวลากรอกข้อมูล:</div>
              <div style="color: #a16207; font-size: 13px;">
                ลิงก์นี้เปิดให้กรอกข้อมูลและจะหมดอายุในวันที่ <strong>{expiry_formatted}</strong> (สามารถบันทึกข้อมูลได้ 1 ครั้ง)
              </div>
            </div>

            <!-- Action Button -->
            <div style="text-align: center; margin: 30px 0;">
              <a href="{portal_url}" style="background-color: #0284c7; color: #ffffff; padding: 14px 32px; text-decoration: none; border-radius: 10px; font-weight: bold; font-size: 14px; display: inline-block; box-shadow: 0 2px 4px rgba(2, 132, 199, 0.2);">
                คลิกเพื่อกรอกวันส่งสินค้า (Supplier Portal)
              </a>
            </div>

            <p style="font-size: 12px; color: #64748b; margin-bottom: 24px;">
              หรือคัดลอกลิงก์ด้านล่างนี้ไปเปิดในเว็บเบราว์เซอร์:<br>
              <a href="{portal_url}" style="color: #0284c7; word-break: break-all; font-size: 11px;">{portal_url}</a>
            </p>

            <!-- Strict NO-REPLY Warning Banner -->
            <div style="background-color: #fef2f2; border: 1px solid #fecaca; border-radius: 10px; padding: 14px 18px; margin-top: 24px;">
              <div style="color: #991b1b; font-size: 12px; font-weight: bold; margin-bottom: 2px;">
                📌 ข้อความอัตโนมัติจากระบบ (No-Reply)
              </div>
              <div style="color: #b91c1c; font-size: 12px; line-height: 1.5;">
                โปรดอย่า Reply หรือตอบกลับอีเมลนี้ เนื่องจากเป็นระบบส่งข้อความอัตโนมัติที่ไม่สามารถรับข้อความตอบกลับได้<br>
                หากมีข้อสงสัยหรือติดปัญหาเรื่องสินค้า กรุณาติดต่อประสานงานฝ่ายจัดซื้อที่ 02-123-1734 ต่อ 311 หรือเจ้าหน้าที่ฝ่ายจัดซื้อโดยตรง
              </div>
            </div>
          </div>

          <!-- Footer -->
          <div style="background-color: #f1f5f9; padding: 16px 32px; border-top: 1px solid #e2e8f0; font-size: 11px; color: #94a3b8; text-align: center;">
            Incoming Raw Material (IRM) System © Window Asia Public Company Limited
          </div>
        </div>
      </body>
    </html>
    """


async def send_single_supplier_email(
    db: AsyncSession,
    supplier: SupplierMaster,
    smtp_config: dict,
    now_dt: datetime,
) -> dict:
    """Send portal link email to a single supplier and lock items under supplier window."""
    if not supplier.email or "@" not in supplier.email:
        return {"status": "skipped", "reason": "No valid email"}

    token_obj = await get_or_create_supplier_token(db, supplier.supplier_code)
    
    # Resolve base URL from SystemSetting 'app_base_url' or environment or default to https://irm.windowasia.com
    stmt_base = select(SystemSetting).where(SystemSetting.key == "app_base_url")
    base_setting = (await db.execute(stmt_base)).scalar_one_or_none()
    
    if base_setting and base_setting.value:
        base_url = base_setting.value.strip().rstrip("/")
    else:
        from app.config import get_settings
        cfg = get_settings()
        base_url = (cfg.FRONTEND_URL or "https://irm.windowasia.com").strip().rstrip("/")

    if "irm.windowasia.com" in base_url and base_url.startswith("http://"):
        base_url = base_url.replace("http://", "https://")
    elif not base_url.startswith("http"):
        base_url = f"https://{base_url}"

    portal_url = f"{base_url}/supplier/portal/{token_obj.token}"
    expiry_formatted = token_obj.expires_at.strftime("%d/%m/%Y เวลา 23:59 น.")

    # Count open PO items for this supplier
    stmt_items = (
        select(POItem)
        .join(POHeader, POItem.po_header_id == POHeader.id)
        .where(POHeader.status == "O", POHeader.supplier_code == supplier.supplier_code)
    )
    po_items = (await db.execute(stmt_items)).scalars().all()
    open_po_count = len(po_items)

    # 1. Update POItem ownership locks
    for item in po_items:
        item.locked_by = "supplier"
        item.lock_expires_at = token_obj.expires_at
        if item.status in ["pending", "estimate"]:
            item.status = "awaiting_supplier"

    # 2. Build email with modern EmailMessage supporting native UTF-8
    import re
    from email.message import EmailMessage

    # Sanitize email in case of accidental Thai keyboard characters (e.g. 'ืn.chaiwat@gmail.com')
    clean_email = re.sub(r"[^\x00-\x7F]+", "", supplier.email).strip()
    if not clean_email or "@" not in clean_email:
        return {"status": "skipped", "reason": "No valid ASCII email format"}

    subject = f"เรียน {supplier.supplier_name} — แจ้งระบุวันส่งมอบวัตถุดิบ (IRM System)"
    html_body = build_no_reply_email_html(supplier.supplier_name, portal_url, expiry_formatted, open_po_count)

    smtp_host = smtp_config.get("smtp_host") or "smtp.gmail.com"
    smtp_port = int(smtp_config.get("smtp_port") or 587)
    smtp_user = smtp_config.get("smtp_user") or "noreply@company.com"
    smtp_pass = smtp_config.get("smtp_password") or ""
    use_tls = smtp_config.get("smtp_use_tls", "true").lower() == "true"
    from_name = smtp_config.get("smtp_from_name") or "IRM System (No-Reply)"

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = formataddr((from_name, smtp_user))
    msg["To"] = clean_email
    msg["Reply-To"] = "noreply@company.com"
    msg.set_content(html_body, subtype="html", charset="utf-8")

    if smtp_port == 465:
        server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=12.0)
    else:
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=12.0)
        if use_tls:
            server.starttls()

    if smtp_pass:
        server.login(smtp_user, smtp_pass)

    server.send_message(msg)
    server.quit()

    supplier.last_sent_at = now_dt
    await db.commit()

    # Log single email
    try:
        from app.services.log_service import record_transaction_log
        await record_transaction_log(
            category="supplier_email",
            action="send_single_email",
            status="success",
            message=f"ส่ง Email แจ้งลิงก์ Portal หา {supplier.supplier_name} ({supplier.email}) สำเร็จ",
            details={
                "supplier_code": supplier.supplier_code,
                "supplier_name": supplier.supplier_name,
                "email": supplier.email,
                "token": token_obj.token,
                "expires_at": token_obj.expires_at.isoformat(),
            },
            records_count=1,
            triggered_by="manual_or_system",
            db=db,
        )
        await db.commit()
    except Exception:
        pass

    return {"status": "success", "supplier_code": supplier.supplier_code, "email": supplier.email}


async def send_batch_portal_emails(db: AsyncSession, max_suppliers: int = 100) -> dict:
    """
    Automated Batch Email Sender with rate limiting:
    - Chunks suppliers into batches of 20
    - Waits 5 seconds between chunks to prevent SMTP provider throttling
    - Sends up to max_suppliers (<= 100) per session
    """
    now_dt = datetime.now(timezone.utc)

    # 1. Fetch SMTP & Batching configs
    keys = [
        "smtp_host", "smtp_port", "smtp_user", "smtp_password", "smtp_use_tls",
        "smtp_from_name", "email_batch_size", "email_batch_delay_seconds", "email_max_per_session",
        "mail_schedule_enabled"
    ]
    settings_rows = (await db.execute(select(SystemSetting).where(SystemSetting.key.in_(keys)))).scalars().all()
    s_map = {s.key: s.value for s in settings_rows}

    # FAIL-SAFE GUARD: Strict lock during implementation phase
    schedule_enabled = s_map.get("mail_schedule_enabled", "false").strip().lower()
    if schedule_enabled not in ("true", "1", "yes"):
        logger.warning("🛡️ [SAFEGUARD] Scheduled Batch Email Broadcast is STRICTLY DISABLED (mail_schedule_enabled=false). Zero emails sent to suppliers during Implementation Phase.")
        return {
            "status": "disabled",
            "message": "Scheduled email broadcast is disabled during Implementation Phase (Safety Lock Active).",
            "total_attempted": 0,
            "sent_count": 0,
            "failed_count": 0,
            "timestamp": now_dt.strftime("%d/%m/%Y %H:%M:%S"),
        }

    batch_size = int(s_map.get("email_batch_size") or 20)
    delay_secs = int(s_map.get("email_batch_delay_seconds") or 5)
    max_session = min(int(s_map.get("email_max_per_session") or 100), max_suppliers)

    # 2. Find eligible suppliers with valid email and active open POs
    stmt_suppliers = (
        select(SupplierMaster)
        .where(SupplierMaster.email.isnot(None))
        .where(SupplierMaster.email != "")
        .where(SupplierMaster.email != "-")
        .order_by(SupplierMaster.supplier_code.asc())
        .limit(max_session)
    )
    suppliers = (await db.execute(stmt_suppliers)).scalars().all()

    sent_count = 0
    failed_count = 0
    errors = []

    logger.info(f"📧 Starting Batch Portal Email Broadcast for {len(suppliers)} suppliers (Batch size: {batch_size}, Delay: {delay_secs}s)...")

    # 3. Chunk and dispatch with sleep intervals
    for i in range(0, len(suppliers), batch_size):
        chunk = suppliers[i : i + batch_size]
        logger.info(f"📧 Sending batch chunk {i // batch_size + 1} ({len(chunk)} suppliers)...")

        for sup in chunk:
            try:
                res = await send_single_supplier_email(db, sup, s_map, now_dt)
                if res["status"] == "success":
                    sent_count += 1
                else:
                    logger.warning(f"Skipped {sup.supplier_code}: {res.get('reason')}")
            except Exception as err:
                failed_count += 1
                err_msg = f"{sup.supplier_code} ({sup.email}): {str(err)}"
                errors.append(err_msg)
                logger.error(f"Failed sending email to {err_msg}")

        # If there are more chunks, wait delay_secs
        if i + batch_size < len(suppliers):
            logger.info(f"⏳ Waiting {delay_secs}s before sending next batch chunk...")
            await asyncio.sleep(delay_secs)

    return {
        "status": "completed",
        "total_attempted": len(suppliers),
        "sent_count": sent_count,
        "failed_count": failed_count,
        "errors": errors[:5],  # Top 5 errors if any
        "timestamp": now_dt.strftime("%d/%m/%Y %H:%M:%S"),
    }


import io
from email.mime.application import MIMEApplication
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.models.user import User


async def generate_pu_remind_excel(db: AsyncSession) -> tuple[bytes, dict]:
    """
    Generates a 2-Sheet Excel workbook for PU User daily reminder:
    - Sheet 1: รายการ PO และ Item ที่ยังไม่ Confirm วันส่งมอบ (Unconfirmed PO Items)
    - Sheet 2: รายการ PO และ Item ที่มีกำหนดส่งของภายในวันนี้ (Today's Scheduled Deliveries)
    """
    from zoneinfo import ZoneInfo
    bkk_tz = ZoneInfo("Asia/Bangkok")
    now_bkk = datetime.now(bkk_tz)
    today_bkk_date = now_bkk.date()

    # Query all open PO Items with Headers and SubItems
    from sqlalchemy.orm import selectinload
    stmt = (
        select(POItem, POHeader)
        .join(POHeader, POItem.po_header_id == POHeader.id)
        .options(selectinload(POItem.sub_items))
        .where(POHeader.status == "O")
        .where(POItem.status != "closed")
        .order_by(POHeader.po_number.desc(), POItem.id.asc())
    )
    rows = (await db.execute(stmt)).all()

    wb = Workbook()

    # Define Styles
    header_fill_blue = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
    header_fill_green = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid") # Emerald
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1: รอ Confirm วันส่งมอบ (Unconfirmed PO Items)
    # ─────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "รอ Confirm วันส่งมอบ"
    ws1.views.sheetView[0].showGridLines = True

    headers_s1 = [
        "#", "PO No.", "PO Date", "Group", "Item Code", "Description",
        "PO Qty", "Unit", "Due Date", "Received Qty", "Remaining Qty",
        "Est. Date", "Est. Qty", "Buyer", "Supplier Code", "Supplier Name", "Status"
    ]
    ws1.append(headers_s1)
    for col_idx in range(1, len(headers_s1) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill_blue
        cell.font = header_font
        cell.alignment = align_center

    unconfirmed_pos = set()
    unconfirmed_items_count = 0
    row_num_s1 = 1

    for item, header in rows:
        if item.status == "confirmed":
            continue

        unconfirmed_items_count += 1
        unconfirmed_pos.add(header.po_number)
        row_num_s1 += 1

        po_dt_str = header.po_date.strftime("%d/%m/%Y") if header.po_date else "-"
        due_dt_str = item.due_date.strftime("%d/%m/%Y") if item.due_date else "-"
        est_dt_str = item.estimate_date.strftime("%d/%m/%Y") if item.estimate_date else "-"
        rem_qty = float(item.remaining_qty if item.remaining_qty is not None else item.quantity)
        est_qty = float(item.estimate_qty if item.estimate_qty is not None else rem_qty)

        status_th = "รอคู่ค้าตอบกลับ" if item.status == "awaiting_supplier" else ("คู่ค้าตอบกลับแล้ว" if item.status == "supplier_responded" else "ยังไม่ระบุวัน")

        row_data = [
            row_num_s1 - 1,
            header.po_number,
            po_dt_str,
            item.item_group or "-",
            item.item_code,
            item.item_name or "-",
            float(item.quantity or 0),
            item.unit or "",
            due_dt_str,
            float(item.received_qty or 0),
            rem_qty,
            est_dt_str,
            est_qty,
            header.buyer_name or "-",
            header.supplier_code or "-",
            header.supplier_name or "-",
            status_th,
        ]
        ws1.append(row_data)

        # Apply cell formatting
        for col_idx in range(1, len(row_data) + 1):
            c = ws1.cell(row=row_num_s1, column=col_idx)
            c.font = data_font
            c.border = thin_border
            if col_idx in (1, 2, 3, 4, 8, 9, 12, 14, 15, 17):
                c.alignment = align_center
            elif col_idx in (7, 10, 11, 13):
                c.alignment = align_right
            else:
                c.alignment = align_left

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 2: กำหนดส่งมอบวันนี้ (Today's Scheduled Deliveries)
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="กำหนดส่งมอบวันนี้")
    ws2.views.sheetView[0].showGridLines = True

    headers_s2 = [
        "#", "PO No.", "PO Date", "Group", "Item Code", "Description",
        "Delivery Date (วันนี้)", "Delivery Qty", "Unit", "Due Date", "Received Qty", "Remaining Qty",
        "Buyer", "Supplier Code", "Supplier Name", "Status"
    ]
    ws2.append(headers_s2)
    for col_idx in range(1, len(headers_s2) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = header_fill_green
        cell.font = header_font
        cell.alignment = align_center

    today_pos = set()
    today_items_count = 0
    row_num_s2 = 1

    for item, header in rows:
        po_dt_str = header.po_date.strftime("%d/%m/%Y") if header.po_date else "-"
        due_dt_str = item.due_date.strftime("%d/%m/%Y") if item.due_date else "-"
        rem_qty = float(item.remaining_qty if item.remaining_qty is not None else item.quantity)
        buyer = header.buyer_name or "-"
        sup_code = header.supplier_code or "-"
        sup_name = header.supplier_name or "-"
        status_label = "Confirmed (ยืนยันแล้ว)" if item.status == "confirmed" else "Estimate (ประมาณการ)"

        # Check sub items
        if item.sub_items:
            for sub in item.sub_items:
                if sub.estimate_date and sub.estimate_date.date() == today_bkk_date:
                    today_items_count += 1
                    today_pos.add(header.po_number)
                    row_num_s2 += 1
                    row_data = [
                        row_num_s2 - 1,
                        header.po_number,
                        po_dt_str,
                        item.item_group or "-",
                        f"↳ {item.item_code}",
                        item.item_name or "-",
                        sub.estimate_date.strftime("%d/%m/%Y"),
                        float(sub.quantity or 0),
                        item.unit or "",
                        due_dt_str,
                        float(item.received_qty or 0),
                        rem_qty,
                        buyer,
                        sup_code,
                        sup_name,
                        status_label,
                    ]
                    ws2.append(row_data)
                    for col_idx in range(1, len(row_data) + 1):
                        c = ws2.cell(row=row_num_s2, column=col_idx)
                        c.font = data_font
                        c.border = thin_border
                        if col_idx in (1, 2, 3, 4, 7, 9, 10, 13, 14, 16):
                            c.alignment = align_center
                        elif col_idx in (8, 11, 12):
                            c.alignment = align_right
                        else:
                            c.alignment = align_left
        elif item.estimate_date and item.estimate_date.date() == today_bkk_date:
            today_items_count += 1
            today_pos.add(header.po_number)
            row_num_s2 += 1
            del_qty = float(item.estimate_qty if item.estimate_qty is not None else rem_qty)
            row_data = [
                row_num_s2 - 1,
                header.po_number,
                po_dt_str,
                item.item_group or "-",
                item.item_code,
                item.item_name or "-",
                item.estimate_date.strftime("%d/%m/%Y"),
                del_qty,
                item.unit or "",
                due_dt_str,
                float(item.received_qty or 0),
                rem_qty,
                buyer,
                sup_code,
                sup_name,
                status_label,
            ]
            ws2.append(row_data)
            for col_idx in range(1, len(row_data) + 1):
                c = ws2.cell(row=row_num_s2, column=col_idx)
                c.font = data_font
                c.border = thin_border
                if col_idx in (1, 2, 3, 4, 7, 9, 10, 13, 14, 16):
                    c.alignment = align_center
                elif col_idx in (8, 11, 12):
                    c.alignment = align_right
                else:
                    c.alignment = align_left

    # Auto adjust column widths for both sheets
    for ws in (ws1, ws2):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()
    excel_buffer.close()

    stats = {
        "unconfirmed_po_count": len(unconfirmed_pos),
        "unconfirmed_item_count": unconfirmed_items_count,
        "today_delivery_po_count": len(today_pos),
        "today_delivery_item_count": today_items_count,
        "date_thai": now_bkk.strftime("%d/%m/%Y"),
        "date_file": now_bkk.strftime("%Y%m%d"),
    }

    return excel_bytes, stats


async def send_pu_daily_reminder_email(
    db: AsyncSession,
    recipient_email: str | None = None,
    triggered_by: str = "scheduler",
) -> dict:
    """
    Dispatches the Daily PU Reminder Email with 2-Sheet Excel attachment.
    Sent to PU Users (or specified test email).
    """
    from zoneinfo import ZoneInfo
    bkk_tz = ZoneInfo("Asia/Bangkok")
    now_bkk = datetime.now(bkk_tz)

    # 1. Fetch SMTP & Config settings
    keys = [
        "smtp_host", "smtp_port", "smtp_user", "smtp_password", "smtp_use_tls",
        "smtp_from_name", "pu_remind_mail_enabled", "pu_remind_mail_time"
    ]
    settings_rows = (await db.execute(select(SystemSetting).where(SystemSetting.key.in_(keys)))).scalars().all()
    s_map = {s.key: s.value for s in settings_rows}

    # If triggered by scheduler, check enable switch
    if triggered_by == "scheduler":
        is_enabled = s_map.get("pu_remind_mail_enabled", "false").strip().lower()
        if is_enabled not in ("true", "1", "yes"):
            logger.info("🛡️ PU Reminder email is disabled (pu_remind_mail_enabled=false). Skipping.")
            return {"status": "skipped", "message": "PU reminder email is disabled in settings"}

    smtp_host = s_map.get("smtp_host", "smtp.gmail.com")
    smtp_port = int(s_map.get("smtp_port") or 587)
    smtp_user = s_map.get("smtp_user", "")
    smtp_pass = s_map.get("smtp_password", "")
    smtp_from = s_map.get("smtp_from_name", "IRM System")
    use_tls = s_map.get("smtp_use_tls", "true").lower() == "true"

    if not smtp_user or not smtp_pass:
        return {"status": "error", "message": "SMTP user or password not configured in System Settings"}

    # 2. Determine Recipients
    recipients = []
    if recipient_email and recipient_email.strip():
        recipients = [recipient_email.strip()]
    else:
        # Fetch active users in 'PU User' group
        stmt_users = (
            select(User)
            .join(User.group)
            .where(User.is_active == True)
            .where(User.email.isnot(None))
            .where(User.email != "")
        )
        pu_users = (await db.execute(stmt_users)).scalars().all()
        recipients = list({u.email.strip() for u in pu_users if u.email and "@" in u.email})

    if not recipients:
        return {"status": "skipped", "message": "No active PU User email addresses found"}

    # 3. Generate 2-Sheet Excel & Summary Statistics
    excel_bytes, stats = await generate_pu_remind_excel(db)

    # 4. Construct HTML Body
    subject = f"IRM System — สรุปรายการติดตามการรับวัตถุดิบและของส่งวันนี้ ({stats['date_thai']})"
    html_body = f"""
    <!DOCTYPE html>
    <html>
      <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
      </head>
      <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #1e293b; background-color: #f8fafc; margin: 0; padding: 24px; line-height: 1.6;">
        <div style="max-width: 650px; margin: 0 auto; background: #ffffff; border-radius: 16px; border: 1px solid #e2e8f0; overflow: hidden; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);">
          <!-- Header -->
          <div style="background: linear-gradient(135deg, #0f172a, #1e3a8a); padding: 24px 32px; color: #ffffff;">
            <h1 style="margin: 0; font-size: 20px; font-weight: bold; letter-spacing: -0.5px;">ระบบติดตามการรับวัตถุดิบ (IRM System)</h1>
            <p style="margin: 4px 0 0 0; font-size: 13px; opacity: 0.9;">สรุปรายการติดตามงานฝ่ายจัดซื้อ ประจำวันที่ {stats['date_thai']}</p>
          </div>

          <!-- Body Content -->
          <div style="padding: 28px 32px;">
            <p style="font-size: 15px; margin-top: 0;">เรียน <strong>ทีมงานฝ่ายจัดซื้อ (Purchasing Team)</strong>,</p>
            <p style="font-size: 14px; color: #475569;">
              ระบบ IRM ขอสรุปภาพรวมรายการสั่งซื้อวัตถุดิบ (PO) ประจำวัน เพื่อการวางแผนและติดตามการส่งมอบ ดังนี้:
            </p>

            <!-- Card 1: Unconfirmed POs -->
            <div style="background-color: #fffbeb; border: 1px solid #fde68a; border-radius: 12px; padding: 16px 20px; margin-bottom: 16px;">
              <div style="color: #92400e; font-size: 14px; font-weight: bold; margin-bottom: 4px;">
                ⚠️ รายการที่ยังไม่ Confirm Delivery Date:
              </div>
              <div style="font-size: 18px; font-weight: bold; color: #b45309;">
                {stats['unconfirmed_po_count']} ใบสั่งซื้อ <span style="font-size: 14px; font-weight: normal; color: #78350f;">({stats['unconfirmed_item_count']:,} รายการสินค้า)</span>
              </div>
              <div style="font-size: 12px; color: #a16207; margin-top: 4px;">
                * ดูรายละเอียดได้ในไฟล์แนบ Sheet ที่ 1 (<code>รอ Confirm วันส่งมอบ</code>)
              </div>
            </div>

            <!-- Card 2: Today's Deliveries -->
            <div style="background-color: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 12px; padding: 16px 20px; margin-bottom: 24px;">
              <div style="color: #166534; font-size: 14px; font-weight: bold; margin-bottom: 4px;">
                🚚 รายการที่มีกำหนดส่งมอบภายในวันนี้:
              </div>
              <div style="font-size: 18px; font-weight: bold; color: #15803d;">
                {stats['today_delivery_po_count']} ใบสั่งซื้อ <span style="font-size: 14px; font-weight: normal; color: #14532d;">({stats['today_delivery_item_count']:,} รายการส่งมอบ)</span>
              </div>
              <div style="font-size: 12px; color: #15803d; margin-top: 4px;">
                * ดูรายละเอียดได้ในไฟล์แนบ Sheet ที่ 2 (<code>กำหนดส่งมอบวันนี้</code>)
              </div>
            </div>

            <!-- Notice & Direct Link -->
            <div style="text-align: center; margin: 28px 0 16px 0;">
              <a href="https://irm.windowasia.com/operation" style="background-color: #1e3a8a; color: #ffffff; padding: 12px 28px; text-decoration: none; border-radius: 10px; font-weight: bold; font-size: 14px; display: inline-block;">
                เปิดเข้าระบบ IRM Operation
              </a>
            </div>

            <p style="font-size: 12px; color: #94a3b8; text-align: center; margin-top: 20px;">
              📎 ไฟล์แนบ: <code>IRM_PU_Daily_Summary_{stats['date_file']}.xlsx</code> (มี 2 Sheet ตามรายละเอียดข้างต้น)
            </p>
          </div>

          <!-- Footer -->
          <div style="background-color: #f1f5f9; padding: 16px 32px; border-top: 1px solid #e2e8f0; font-size: 11px; color: #64748b; text-align: center;">
            อีเมลฉบับนี้ส่งโดยระบบอัตโนมัติ IRM (Incoming Raw Material System) — Window Asia PCL.
          </div>
        </div>
      </body>
    </html>
    """

    # 5. Build and Send Email
    file_name = f"IRM_PU_Daily_Summary_{stats['date_file']}.xlsx"
    sent_recipients = []
    errors = []

    for recipient in recipients:
        try:
            msg = MIMEMultipart()
            msg["Subject"] = subject
            msg["From"] = formataddr((smtp_from, smtp_user))
            msg["To"] = recipient

            # Attach HTML part
            msg.attach(MIMEText(html_body, "html", "utf-8"))

            # Attach Excel
            part = MIMEApplication(excel_bytes, _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.add_header("Content-Disposition", "attachment", filename=file_name)
            msg.attach(part)

            # Send via SMTP
            await asyncio.to_thread(_send_smtp_sync, smtp_host, smtp_port, smtp_user, smtp_pass, use_tls, msg, [recipient])
            sent_recipients.append(recipient)
        except Exception as e:
            err_msg = f"Failed sending PU remind email to {recipient}: {e}"
            logger.error(err_msg)
            errors.append(err_msg)

    # Log in transaction_logs
    try:
        await record_transaction_log(
            category="pu_remind_email",
            action="send_daily_remind",
            status="SUCCESS" if sent_recipients else "FAILED",
            message=f"ส่งอีเมลสรุปงานให้จัดซื้อสำเร็จ {len(sent_recipients)} ท่าน (Unconfirmed: {stats['unconfirmed_item_count']} รายการ, Today: {stats['today_delivery_item_count']} รายการ)",
            details=f"Recipients: {', '.join(sent_recipients)} | Trigger: {triggered_by} | Errors: {errors}",
            db=db,
        )
        await db.commit()
    except Exception:
        pass

    return {
        "status": "success" if sent_recipients else "failed",
        "sent_count": len(sent_recipients),
        "recipients": sent_recipients,
        "errors": errors,
        "stats": stats,
    }

